from Tank import Tank
from Valve import Valve
from Pump import Pump
from PumpState import PumpState
from ValveState import ValveState
from TankState import TankState
import openpyxl
from graphviz import Graph
from graphviz import Digraph


def is_stable(item_list: list, type):
    if type == "tank":
        for item in item_list:
            if not (item.state.value == 0 or item.state.value == 1 or item.state.value == 2):
                return False
    else:
        for item in item_list:
            if not (item.state.value == 1 or item.state.value == 2):
                return False
    return True


def is_same(path1, path2):
    if path1[0] != path2[0] or path1[2] != path2[2]:
        return False
    if len(path1[1]) != len(path2[1]):
        return False
    for item in path1[1]:
        if item not in path2[1]:
            return False
    return True


def add_path(path_set, path):
    for p in path_set:
        if is_same(path, p):
            return
    path_set.append(path)


def find_intersection(a, b):
    c = []
    for i in a:
        if i in b:
            c.append(i)
    return c


def arrange_order(valve_pump_list: list, Pump_list):
    for item in valve_pump_list:
        if item in Pump_list:
            valve_pump_list.remove(item)
            valve_pump_list.append(item)


def find_path(path_set, s, t):
    for path in path_set:
        if path[0] == s and path[2] == t:
            return path[1]


data_file_path = "食品饮料流程工艺数据.xlsx"
data_workbook = openpyxl.load_workbook(data_file_path)

sheet = data_workbook[data_workbook.sheetnames[0]]

window_size = 4  # 滑动窗口大小

g = Graph("Structure Graph")

Tank_list = [Tank(x[1].value[:4], window_size, g) for x in sheet.iter_cols(min_col=3, max_col=8)]
Valve_list = [Valve(x[1].value[:4], window_size, g) for x in sheet.iter_cols(min_col=9, max_col=28)]
Pump_list = [Pump(x[1].value[:6], window_size, g) for x in sheet.iter_cols(min_col=29, max_col=30)]

outside = Tank("outside", window_size, g, False)

opened_valve_pump = []
tank_importing_list = []
tank_exporting_list = []
tank_state = []
valve_state = []
path_set = []
path = []

# 开始检测
for row in sheet.iter_rows(min_row=3):
    for idx, cell in enumerate(row[2:8]):
        Tank_list[idx].renew_amount(cell.value)
    for idx, cell in enumerate(row[8:28]):
        Valve_list[idx].renew_state(cell.value)
    for idx, cell in enumerate(row[28:30]):
        Pump_list[idx].renew_state(cell.value)

    for valve in Valve_list:
        if valve.state == ValveState.open and valve not in opened_valve_pump:
            opened_valve_pump.append(valve)
        if valve.state != ValveState.open and valve in opened_valve_pump:
            opened_valve_pump.remove(valve)

    for pump in Pump_list:
        if pump.state == PumpState.open and pump not in opened_valve_pump:
            opened_valve_pump.append(pump)
        if pump.state != PumpState.open and pump in opened_valve_pump:
            opened_valve_pump.remove(pump)

    if is_stable(Valve_list, "valve") and is_stable(Pump_list, "pump"):
        for tank in Tank_list:
            if tank.state == TankState.exporting:
                if tank not in tank_exporting_list:
                    tank_exporting_list.append(tank)
            elif tank.state == TankState.importing:
                if tank not in tank_importing_list:
                    tank_importing_list.append(tank)
            else:
                if tank in tank_exporting_list:
                    tank_exporting_list.remove(tank)
                if tank in tank_importing_list:
                    tank_importing_list.remove(tank)

        # if len(tank_state) != 0:
        #     remove_repeat_path(tank_changing_list, opened_valve, tank_state[-1], valve_state[-1])

        if len(tank_state) != 0:
            if (tank_exporting_list != tank_state[-1][0] or tank_importing_list != tank_state[-1][1]
                or opened_valve_pump != valve_state[-1]) and opened_valve_pump != []:

                new_opened_valve_pump = opened_valve_pump.copy()
                if len(valve_state) != 0 and new_opened_valve_pump != valve_state[-1] and len(
                        new_opened_valve_pump) >= len(valve_state[-1]) \
                        and new_opened_valve_pump[:len(valve_state[-1])] == valve_state[-1]:
                    new_opened_valve_pump = new_opened_valve_pump[len(valve_state[-1]):]

                new_tank_exporting_list = tank_exporting_list.copy()
                new_tank_importing_list = tank_importing_list.copy()

                if len(new_tank_exporting_list) == 0 or len(new_tank_importing_list) > len(new_tank_exporting_list):
                    new_tank_exporting_list.append(outside)
                if len(new_tank_importing_list) == 0 or len(new_tank_importing_list) < len(new_tank_exporting_list):
                    new_tank_importing_list.append(outside)

                add_path(path_set,
                         [new_tank_exporting_list[-1], new_opened_valve_pump.copy(), new_tank_importing_list[-1],
                          new_tank_exporting_list[-1].get_volume(), new_tank_importing_list[-1].get_volume()])
                # print([new_tank_exporting_list[-1], new_tank_exporting_list[-1].get_volume(), new_tank_importing_list[-1]])

                valve_state.append(opened_valve_pump.copy())
                tank_state.append([tank_exporting_list.copy(), tank_importing_list.copy()])
        else:
            if tank_exporting_list != [] or tank_importing_list != [] or opened_valve_pump != []:
                new_tank_exporting_list = tank_exporting_list.copy()
                new_tank_importing_list = tank_importing_list.copy()

                if len(new_tank_exporting_list) == 0 or len(new_tank_importing_list) > len(new_tank_exporting_list):
                    new_tank_exporting_list.append(outside)
                if len(new_tank_importing_list) == 0 or len(new_tank_importing_list) < len(new_tank_exporting_list):
                    new_tank_importing_list.append(outside)

                valve_state.append(opened_valve_pump.copy())
                tank_state.append([tank_exporting_list.copy(), tank_importing_list.copy()])
                add_path(path_set, [new_tank_exporting_list[-1], opened_valve_pump.copy(), new_tank_importing_list[-1],
                                    new_tank_exporting_list[-1].get_volume(), new_tank_importing_list[-1].get_volume()])
                # print([new_tank_exporting_list[-1], new_tank_exporting_list[-1].get_volume(), new_tank_importing_list[-1]])

# 找出的总阀门
tank_out = dict()
tank_in = dict()
for path in path_set:
    if path[0] != outside:
        if path[0] not in tank_out:
            tank_out[path[0]] = path[1].copy()
        else:
            tank_out[path[0]] = find_intersection(tank_out[path[0]], path[1])

    if path[2] not in tank_in:
        tank_in[path[2]] = path[1].copy()
    else:
        tank_in[path[2]] = find_intersection(tank_in[path[2]], path[1])

for i in tank_out:
    arrange_order(tank_out[i], Pump_list)
    i.connect(tank_out[i][0])
    for j in range(len(tank_out[i]) - 1):
        tank_out[i][j].connect(tank_out[i][j + 1])

    for k in tank_in:
        p = tank_out[i][-1]
        path = find_path(path_set, i, k)
        for item in path:
            if item not in tank_out[i] and item not in tank_in[k]:
                p.connect(item)
                p = item
        for item in tank_in[k]:
            p.connect(item)
            p = item
        if k != outside:
            p.connect(k)

for path in path_set:
    if path[0] == outside and path[2] == outside:
        p = None
        for item in path[1]:
            if isinstance(item, Pump):
                p = item

        for item in path[1]:
            if not isinstance(item, Pump):
                if item not in p.next:
                    item.connect(p)

# print(tank_out)
# print(tank_in)

g.view()

s = Digraph('State Machine')

last_state_name = 'export = outside\nimport = outside'
s.node(name=last_state_name, shape='box')

for path in path_set:
    state_name = 'export = {}\nimport = {}'.format(path[0].get_name(), path[2].get_name())
    if path[0].get_name() == "outside":
        condition1 = ''
    else:
        condition1 = "volume of " + path[0].get_name() + "=" + str(path[-2])
    if path[2].get_name() == "outside":
        condition2 = ''
    else:
        condition2 = "volume of " + path[2].get_name() + "=" + str(path[-1])
    s.node(name=state_name, shape='box')
    s.edge(tail_name=last_state_name, head_name=state_name, label=condition1+'\n'+condition2, arrowhead='vee')
    last_state_name = state_name
s.node(name='Start', color='red')
s.edge('Start', 'export = outside\nimport = outside', label='TA01 = 1000\nTA02 = 1000', arrowhead='vee')
s.node(name='End', color='red')
s.edge('export = outside\nimport = outside', 'End', arrowhead='vee')
s.view()

for path in path_set:
    print(path)